from django.http import HttpResponse
from upload.forms import UploadFileForm
from django.shortcuts import render
from upload.models import *
import openpyxl


def upload_file(request):
    def get_last_used_row(sheet):
        row = sheet.max_row
        while row > 0:
            cells = sheet[row]
            if all([cell.value is None for cell in cells]):
                row -= 1
            else:
                break
        if row == 0:
            return 0, 0
        return row

    def get_first_row_by_value(sheet, value):
        for row in sheet.iter_rows(min_row=1, min_col=1, max_row=get_last_used_row(wb.worksheets[0]), max_col=1):
            for cell in row:
                if cell.value == value:
                    return sheet.cell(row=cell.row, column=cell.column).row

    def get_nonempty_rows_in_first_column(sheet):
        filled = 0
        for row in range(get_first_row_by_value(wb.worksheets[0], 1), get_last_used_row(wb.worksheets[0]) + 1):
            if sheet.cell(row=row, column=1).value is not None:
                filled += 1
        return filled


    if request.method == 'POST':
        upload_file_from = UploadFileForm(request.POST, request.FILES)
        if upload_file_from.is_valid():
            file = request.FILES['file']
            # >>>
            wb = openpyxl.load_workbook(file)
            sheet = wb.active
            testName = sheet.cell(row=1, column=1).value
            testDescriptionShort = sheet.cell(row=2, column=1).value
            testDescriptionLong = sheet.cell(row=3, column=1).value
            QuestionsArr = []
            Answers = [[0 for j in range(0)] for i in range(get_nonempty_rows_in_first_column(wb.worksheets[0]))]
            Corrects = [[0 for j in range(0)] for i in range(get_nonempty_rows_in_first_column(wb.worksheets[0]))]

            for row in range(get_first_row_by_value(wb.worksheets[0], 1), get_last_used_row(wb.worksheets[0]) + 1):
                if sheet.cell(row, 1).value is not None:
                    QuestionsArr.append(sheet.cell(row, 2).value)
                else:
                    Answers[len(QuestionsArr) - 1].append(sheet.cell(row, 2).value)
                    if sheet.cell(row, 3).value == 0:
                        Corrects[len(QuestionsArr) - 1].append(False)
                    if sheet.cell(row, 3).value == 1:
                        Corrects[len(QuestionsArr) - 1].append(True)

            db_testName = tests(name = testName, description=testName)
            db_testName.save()
            for Question in range(len(QuestionsArr)):
                answer = Question
                db_questions = questions(test_name = testName, name = QuestionsArr[Question], type = 'NONE', order= Question)
                db_questions.save()
                for answer in range(len(Answers[Question])):
                    db_answer = answers(questions_name = QuestionsArr[Question], answer = Answers[Question][answer], correct = Corrects[Question][answer], comments = '', order = answer)
                    db_answer.save()

            return HttpResponse('Тест успешно загружен', status=200)
            # >>>
    else:
        upload_file_from = UploadFileForm()

    context = {
        'form': upload_file_from
    }
    return render(request, 'upload/upload.html', context=context)